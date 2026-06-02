from __future__ import annotations
from pathlib import Path
from datetime import datetime
from .utils import clean_text, to_decimal


def _excel_date(v):
    try:
        import xlrd
        if isinstance(v, (int, float)) and v > 20000:
            dt = xlrd.xldate.xldate_as_datetime(v, 0)
            return dt.strftime('%Y%m%d')
    except Exception:
        pass
    return clean_text(v)


def parse_excel(path: str) -> tuple[dict, list[dict]]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)
    suffix = p.suffix.lower()
    if suffix == '.xls':
        return _parse_xls(str(p))
    if suffix in ('.xlsx', '.xlsm'):
        return _parse_xlsx(str(p))
    raise ValueError('صيغة الملف غير مدعومة. استخدم XLS أو XLSX')


def _parse_xls(path: str):
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    def cell(r,c):
        if r < sh.nrows and c < sh.ncols:
            return sh.cell_value(r,c)
        return ''
    settings={}
    rows=[]
    # Rajhi Wage Details fixed structure from client's file
    if sh.nrows > 12 and 'Header Section' in str(cell(3,0)):
        settings = {
            'establishment_name': clean_text(cell(2,0)),
            'establishment_bank': clean_text(cell(7,1)) or 'RJHI',
            'establishment_id': clean_text(cell(7,2)),
            'account_number': clean_text(cell(7,3)),
            'currency': clean_text(cell(7,4)) or 'SAR',
            'value_date': _excel_date(cell(7,5)),
            'debit_date': _excel_date(cell(7,7)),
            'file_reference_prefix': clean_text(cell(7,8)) or 'WPS',
            'mol_establishment_id': clean_text(cell(7,10)),
        }
        for r in range(12, sh.nrows):
            net = cell(r,1); iban=clean_text(cell(r,2)); name=clean_text(cell(r,3))
            if not iban or not name:
                continue
            rows.append({
                'name': name,
                'iban': iban,
                'gov_id': clean_text(cell(r,11)),
                'nationality': '',
                'worker_type': 'غير سعودي',
                'basic_salary': str(to_decimal(cell(r,7))),
                'housing_allowance': str(to_decimal(cell(r,8))),
                'other_earnings': str(to_decimal(cell(r,9))),
                'deductions': str(to_decimal(cell(r,10))),
            })
        return settings, rows
    # Generic fallback first row headers
    headers=[clean_text(cell(0,c)).lower() for c in range(sh.ncols)]
    for r in range(1, sh.nrows):
        values={headers[c]: cell(r,c) for c in range(len(headers)) if headers[c]}
        row=_map_generic(values)
        if row.get('name') and row.get('iban'):
            rows.append(row)
    return settings, rows


def _parse_xlsx(path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    def cell(r,c):
        return ws.cell(r+1,c+1).value
    rows=[]; settings={}
    if ws.max_row > 12 and 'Header Section' in str(cell(3,0)):
        settings={
            'establishment_name': clean_text(cell(2,0)),
            'establishment_bank': clean_text(cell(7,1)) or 'RJHI',
            'establishment_id': clean_text(cell(7,2)),
            'account_number': clean_text(cell(7,3)),
            'currency': clean_text(cell(7,4)) or 'SAR',
            'value_date': _excel_date(cell(7,5)),
            'debit_date': _excel_date(cell(7,7)),
            'file_reference_prefix': clean_text(cell(7,8)) or 'WPS',
            'mol_establishment_id': clean_text(cell(7,10)),
        }
        for r in range(12, ws.max_row):
            iban=clean_text(cell(r,2)); name=clean_text(cell(r,3))
            if not iban or not name: continue
            rows.append({
                'name': name, 'iban': iban, 'gov_id': clean_text(cell(r,11)),
                'nationality': '', 'worker_type': 'غير سعودي',
                'basic_salary': str(to_decimal(cell(r,7))), 'housing_allowance': str(to_decimal(cell(r,8))),
                'other_earnings': str(to_decimal(cell(r,9))), 'deductions': str(to_decimal(cell(r,10))),
            })
        return settings, rows
    headers=[clean_text(cell(0,c)).lower() for c in range(ws.max_column)]
    for r in range(1, ws.max_row):
        values={headers[c]: cell(r,c) for c in range(len(headers)) if headers[c]}
        row=_map_generic(values)
        if row.get('name') and row.get('iban'):
            rows.append(row)
    return settings, rows


def _get(values, names):
    for n in names:
        for k,v in values.items():
            if n in k:
                return v
    return ''


def _map_generic(values: dict):
    return {
        'name': clean_text(_get(values,['name','اسم','العامل','employee'])),
        'gov_id': clean_text(_get(values,['id','هوية','اقامة','إقامة','government'])),
        'iban': clean_text(_get(values,['iban','آيبان','ايبان','account'])),
        'nationality': clean_text(_get(values,['nationality','جنسية'])),
        'worker_type': clean_text(_get(values,['type','سعودي'])) or 'غير سعودي',
        'basic_salary': str(to_decimal(_get(values,['basic','راتب','salary']))),
        'housing_allowance': str(to_decimal(_get(values,['housing','سكن']))),
        'other_earnings': str(to_decimal(_get(values,['other','بدلات','allowance']))),
        'deductions': str(to_decimal(_get(values,['deduction','خصم','خصومات']))),
    }
